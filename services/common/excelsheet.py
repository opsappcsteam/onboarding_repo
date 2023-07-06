from openpyxl import load_workbook
import os

def excel_path():
    excel_folder_path = './excels/'
    for file in os.listdir(excel_folder_path):
        if os.path.isfile(os.path.join(excel_folder_path, file)) and '.xlsx' in file:
            excel_file_path = excel_folder_path + file
            return excel_file_path

def excel_name():
    excel_name = excel_path().split('/')[-1].split('.xlsx')[0]
    return excel_name

def excel_sheets():
    workbook = load_workbook(excel_path())
    sheetlist = workbook.sheetnames
    exclude_sheets = ['Guide', 'FAQ', 'Changelog', 'Project Details']
    for sheet in exclude_sheets:
        if sheet in sheetlist:
            sheetlist.remove(sheet)
    if sheetlist == []:
        return '🚫 No Services Detected'
    return sheetlist