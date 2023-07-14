from openpyxl import Workbook

def onboard_csv_content(operation_array, project_name, role_array, email_array):
    if 'onboard' in operation_array:
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = 'Onboard'
        sheet1.append(['operation', 'username', 'projectName', 'role', 'email'])
        sheet1.append(['ONBOARD', project_name, project_name, 'System User', 'DF_ECS@dsta.gov.sg'])
        i = 0
        while i < len(operation_array):
            if operation_array[i] == 'Onboard':
                sheet1.append(['ONBOARD', email_array[i], project_name, role_array[i], email_array[i]])
            else:
                pass
            i += 1
        workbook.save(f'./artifacts/onboarding/{project_name}-Survey_Onboard.csv')
        print('Survey Onboard: Generated')
    else:
        print('Survey Onboard: Skipped')

def reset_csv_content(operation_array, project_name, role_array, email_array):
    if 'reset' in operation_array:
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = 'Reset'
        sheet1.append(['operation', 'username', 'projectName', 'role', 'email'])

        i = 0
        while i < len(operation_array):
            if operation_array[i] == 'Reset':
                sheet1.append(['RESET', email_array[i], project_name, role_array[i], email_array[i]])
            else:
                pass
            i += 1
        workbook.save(f'./artifacts/onboarding/{project_name}-Survey_Reset.csv')
        print('Survey Reset: Generated')
    else:
        print('Survey Reset: Skipped')

def unlock_csv_content(operation_array, project_name, role_array, email_array):
    if 'unlock' in operation_array:
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = 'Unlock'
        sheet1.append(['operation', 'username', 'projectName', 'role', 'email'])

        i = 0
        while i < len(operation_array):
            if operation_array[i] == 'Unlock':
                sheet1.append(['UNLOCK', email_array[i], project_name, role_array[i], email_array[i]])
            else:
                pass
            i += 1
        workbook.save(f'./artifacts/onboarding/{project_name}-Survey_Unlock.csv')
        print('Survey Unlock: Generated')
    else:
        print('Survey Unlock: Skipped')
